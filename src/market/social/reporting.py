"""Reporting engine: PDF/CSV/Excel exports (pustaka/78).

Provides:
- Report generation (portfolio summary, performance, tax)
- CSV export
- Excel export (openpyxl)
- PDF export (reportlab)
- Report scheduling and templates
"""

from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass, field
from datetime import UTC, datetime
from enum import Enum
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


class ReportFormat(Enum):
    """Supported report formats."""

    CSV = "csv"
    EXCEL = "excel"
    PDF = "pdf"
    JSON = "json"


class ReportType(Enum):
    """Types of reports."""

    PORTFOLIO_SUMMARY = "portfolio_summary"
    PERFORMANCE = "performance"
    TRADE_HISTORY = "trade_history"
    TAX_SUMMARY = "tax_summary"
    RISK_ANALYSIS = "risk_analysis"
    CUSTOM = "custom"


@dataclass
class ReportConfig:
    """Configuration for a report."""

    report_type: ReportType
    title: str = ""
    format: ReportFormat = ReportFormat.CSV
    include_charts: bool = True
    include_summary: bool = True
    date_from: str = ""
    date_to: str = ""
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass
class ReportResult:
    """Result of a report generation."""

    report_id: str
    report_type: ReportType
    format: ReportFormat
    title: str
    file_path: str | None = None
    content: bytes | None = None
    rows: int = 0
    generated_at: str = field(default_factory=lambda: datetime.now(UTC).isoformat())
    error: str = ""


class ReportingEngine:
    """Report generation and export engine.

    Supports CSV, Excel (openpyxl), and PDF (reportlab) formats.
    Falls back gracefully if optional libraries are not installed.
    """

    def __init__(self, output_dir: Path | str | None = None) -> None:
        self._output_dir = Path(output_dir) if output_dir else Path("/tmp/market_reports")
        self._report_counter = 0

    def generate_csv(
        self,
        data: list[dict[str, Any]],
        title: str,
        filename: str | None = None,
    ) -> ReportResult:
        """Generate a CSV report.

        Args:
            data: List of dicts to export.
            title: Report title.
            filename: Optional filename (without extension).

        Returns:
            ReportResult with file path or content.
        """
        self._report_counter += 1
        report_id = f"RPT-{self._report_counter:05d}"

        if not data:
            return ReportResult(
                report_id=report_id,
                report_type=ReportType.CUSTOM,
                format=ReportFormat.CSV,
                title=title,
                error="No data to export",
            )

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)

        content = output.getvalue().encode()

        file_path = None
        if filename:
            self._output_dir.mkdir(parents=True, exist_ok=True)
            file_path = str(self._output_dir / f"{filename}.csv")
            Path(file_path).write_bytes(content)

        return ReportResult(
            report_id=report_id,
            report_type=ReportType.CUSTOM,
            format=ReportFormat.CSV,
            title=title,
            file_path=file_path,
            content=content,
            rows=len(data),
        )

    def generate_excel(
        self,
        data: list[dict[str, Any]],
        title: str,
        filename: str | None = None,
        sheet_name: str = "Report",
    ) -> ReportResult:
        """Generate an Excel report.

        Args:
            data: List of dicts to export.
            title: Report title.
            filename: Optional filename.
            sheet_name: Excel sheet name.

        Returns:
            ReportResult with file path or content.
        """
        self._report_counter += 1
        report_id = f"RPT-{self._report_counter:05d}"

        if not data:
            return ReportResult(
                report_id=report_id,
                report_type=ReportType.CUSTOM,
                format=ReportFormat.EXCEL,
                title=title,
                error="No data to export",
            )

        try:
            from openpyxl import Workbook  # type: ignore[import-untyped]

            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Write headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

            # Auto-fit column widths (approximate)
            for col_idx, header in enumerate(headers, 1):
                max_len = max(
                    len(str(header)),
                    max(len(str(row.get(header, ""))) for row in data),
                )
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

            buf = io.BytesIO()
            wb.save(buf)
            content = buf.getvalue()

            file_path = None
            if filename:
                self._output_dir.mkdir(parents=True, exist_ok=True)
                file_path = str(self._output_dir / f"{filename}.xlsx")
                Path(file_path).write_bytes(content)

            return ReportResult(
                report_id=report_id,
                report_type=ReportType.CUSTOM,
                format=ReportFormat.EXCEL,
                title=title,
                file_path=file_path,
                content=content,
                rows=len(data),
            )

        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV")
            return self.generate_csv(data, title, filename)

    def generate_pdf(
        self,
        data: list[dict[str, Any]],
        title: str,
        filename: str | None = None,
        summary: str = "",
    ) -> ReportResult:
        """Generate a PDF report.

        Args:
            data: List of dicts to export.
            title: Report title.
            filename: Optional filename.
            summary: Optional summary text.

        Returns:
            ReportResult with file path or content.
        """
        self._report_counter += 1
        report_id = f"RPT-{self._report_counter:05d}"

        if not data:
            return ReportResult(
                report_id=report_id,
                report_type=ReportType.CUSTOM,
                format=ReportFormat.PDF,
                title=title,
                error="No data to export",
            )

        try:
            from reportlab.lib import colors  # type: ignore[import-untyped]
            from reportlab.lib.pagesizes import A4  # type: ignore[import-untyped]
            from reportlab.lib.styles import (  # type: ignore[import-untyped]
                getSampleStyleSheet,
            )
            from reportlab.platypus import (  # type: ignore[import-untyped]
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
            )

            buf = io.BytesIO()
            doc = SimpleDocTemplate(buf, pagesize=A4)
            styles = getSampleStyleSheet()
            story: list[Any] = []

            # Title
            story.append(Paragraph(title, styles["Title"]))
            story.append(Spacer(1, 12))

            # Summary
            if summary:
                story.append(Paragraph(summary, styles["Normal"]))
                story.append(Spacer(1, 12))

            # Table
            headers = list(data[0].keys())
            table_data = [headers]
            for row in data:
                table_data.append([str(row.get(h, "")) for h in headers])

            table = Table(table_data)
            table.setStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("FONTSIZE", (0, 1), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ])
            story.append(table)

            doc.build(story)
            content = buf.getvalue()

            file_path = None
            if filename:
                self._output_dir.mkdir(parents=True, exist_ok=True)
                file_path = str(self._output_dir / f"{filename}.pdf")
                Path(file_path).write_bytes(content)

            return ReportResult(
                report_id=report_id,
                report_type=ReportType.CUSTOM,
                format=ReportFormat.PDF,
                title=title,
                file_path=file_path,
                content=content,
                rows=len(data),
            )

        except ImportError:
            logger.warning("reportlab not installed, falling back to CSV")
            return self.generate_csv(data, title, filename)

    def generate(
        self,
        data: list[dict[str, Any]],
        config: ReportConfig,
        filename: str | None = None,
    ) -> ReportResult:
        """Generate a report in the specified format.

        Args:
            data: Report data.
            config: Report configuration.
            filename: Optional filename.

        Returns:
            ReportResult.
        """
        title = config.title or config.report_type.value.replace("_", " ").title()

        if config.format == ReportFormat.CSV:
            return self.generate_csv(data, title, filename)
        elif config.format == ReportFormat.EXCEL:
            return self.generate_excel(data, title, filename)
        elif config.format == ReportFormat.PDF:
            return self.generate_pdf(data, title, filename)
        elif config.format == ReportFormat.JSON:
            import json

            self._report_counter += 1
            report_id = f"RPT-{self._report_counter:05d}"
            content = json.dumps(data, indent=2, default=str).encode()
            return ReportResult(
                report_id=report_id,
                report_type=config.report_type,
                format=ReportFormat.JSON,
                title=title,
                content=content,
                rows=len(data),
            )

        return ReportResult(
            report_id="RPT-ERR",
            report_type=config.report_type,
            format=config.format,
            title=title,
            error=f"Unsupported format: {config.format}",
        )

    def portfolio_summary_report(
        self,
        holdings: list[dict[str, Any]],
        format: ReportFormat = ReportFormat.CSV,
    ) -> ReportResult:
        """Generate a portfolio summary report.

        Args:
            holdings: List of holding dicts.
            format: Output format.

        Returns:
            ReportResult.
        """
        config = ReportConfig(
            report_type=ReportType.PORTFOLIO_SUMMARY,
            title="Portfolio Summary Report",
            format=format,
        )
        return self.generate(holdings, config, filename="portfolio_summary")

    def performance_report(
        self,
        performance_data: list[dict[str, Any]],
        format: ReportFormat = ReportFormat.CSV,
    ) -> ReportResult:
        """Generate a performance report.

        Args:
            performance_data: Performance data rows.
            format: Output format.

        Returns:
            ReportResult.
        """
        config = ReportConfig(
            report_type=ReportType.PERFORMANCE,
            title="Performance Report",
            format=format,
        )
        return self.generate(performance_data, config, filename="performance")
